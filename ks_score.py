# -*- coding: utf-8 -*-
"""
Created on Wed May23 23:25 2018

@author: Yifan Zhang
"""
import os
import numpy as np
import pandas as pd
from sklearn.metrics import roc_curve
from openpyxl import load_workbook
from openpyxl.chart import (
      LineChart,
      Reference,
  )

score = ['scoreautofin',
       'scoreautolea', 'scorecashoff', 'scoreconsoff', 'scoreconson',
       'scorecreditbt', 'scorelargecashv2', 'scorepettycashv1',
       'scorerevoloan']
point = np.arange(300,1001,50)

def load_data(option):
	'''
	Function: tune the process of loading big data
	input: 	url: filepath
			option: dict type
	'''
	### extract parameters 
	url = option['url']
	lists = option['lists']
	cus_type = option['cus_type']
	y = option['y']
	###  optimized the memory
	dtype =  dict(zip(score,[np.float16]*9))
	dtype[cus_type] = 'category'
	dtype[y] = np.float16
	f = open(url,encoding='utf-8')
	scoredf = pd.read_csv(f,engine = 'c',skiprows =[1],usecols =score+lists+['flag_score'],dtype = dtype,memory_map = True)
	### load the data and drop the row containing chinese character and fillna 
	scoredf = scoredf[scoredf[y]!=0.5]
	return scoredf

def score_distribution(df,cus_type,point):
    writer = pd.ExcelWriter('score_distribution.xlsx')
    score_distribution = {}
    for i in list(df[cus_type].unique()):
        for j in score:
            score_distribution[i + '_' + j] = pd.cut(df[df[cus_type] == i][str(j)], point, include_lowest=True).value_counts(sort=False)
            score_distribution_df = pd.DataFrame(score_distribution)
    score_distribution_df.index = score_distribution_df.index.astype(str)
    score_distribution_df.to_excel(writer,'score_distribution.xlsx')
    writer.save()

def ks_cus_type(df,cus_type,y):
	'''
	input: df : dataframe
	       cus_type: type str, the column name of cus_type
	       y: type str, the column name of y
	output: csvfile 'ks_result.csv' 
	'''
	index = list(df[cus_type].unique())+['all']
	res = pd.DataFrame(index =index,columns=score)
	for col1 in index:
	    for col2 in score: 
	        if col1 == 'all':
	            index = df[col2].dropna().index
	            print('1')
	            y_train = df.loc[index][y]
	            y_preds_train_1 = df.loc[index][col2].astype(int)
	        else:
	            index = df[df[cus_type]==col1][col2].dropna().index
	            y_train = df[df[cus_type]==col1].loc[index][y]
	            y_preds_train_1 = df[df[cus_type]==col1].loc[index][col2].astype(int)
	        fpr, tpr, _ = roc_curve(y_true=y_train, y_score=y_preds_train_1)
	        res.loc[col1][col2] = max(fpr-tpr)
	        #print('训练集ks:', max(fpr-tpr))
	res.to_csv('ks_result.csv')

def cus_score(df,cus_type,point):
	'''
	function: count every cus_type's nine score at differet bins.
	input: df : dataframe
	       cus_type: type str, the column name of cus_type
	       point: gloval variable to set bins 
	output: excel file 'cus_score.xlsx' 
	'''
	writer = pd.ExcelWriter('cus_score.xlsx')
	columns1 = score
	columns2 = list(df[cus_type].unique())
	for col1 in columns2:
	    df1 = pd.DataFrame()
	    df2 = pd.DataFrame()
	    for col2 in columns1:
	        index =  df[df[cus_type]==col1][col2].dropna().index
	        df1[col2] = pd.cut(df[df[cus_type] == col1].loc[index][col2].astype(int), point, include_lowest=True).value_counts(sort=False)
	        df2[col2+'_distri'] = df1[col2]/df1[col2].sum()
	    res= pd.concat([df1, df2], axis=1)
	    res.index = res.index.astype(str)
	    res.to_excel(writer,col1)
	writer.save()

def score_cus(df,cus_type,point):
	'''
	function: count every score's distribution for every cus_type at differet bins.
	input: df : dataframe
	       cus_type: type str, the column name of cus_type
	       point: gloval variable to set bins
	output: excel file 'score_cus.xlsx' 
	'''
	writer = pd.ExcelWriter('score_cus.xlsx')
	columns1 = score
	columns2 = list(df[cus_type].unique())
	for col2 in columns1:
	    df1 = pd.DataFrame()
	    df2 = pd.DataFrame()
	    for col1 in columns2:
	        index =  df[df[cus_type]==col1][col2].dropna().index
	        df1[col1] = pd.cut(df[df[cus_type] == col1].loc[index][col2].astype(int), point, include_lowest=True).value_counts(sort=False)
	        df2[col1+'_distri'] = df1[col1]/df1[col1].sum()
	    res= pd.concat([df1, df2], axis=1)
	    res.index = res.index.astype(str)
	    res.to_excel(writer,col2)
	writer.save()

def insert_plot(url,start_row,y_title=None,x_title=None):
	'''
	function: insert chart into every function
	input: 	url:  filepath
			start_row: int, the number of lines you need 
	output: excel file
	'''
  colors = ['FF6699','FF0033','FFFF33','FF6600','00CC00','330066','3399FF','663399','FF6699','9966FF','99CC00']
  wb = load_workbook(url)
  sheetnames = wb.get_sheet_names()
  for sheetname in sheetnames:
    ws = wb[sheetname]
    c1 = LineChart()
    c1.title = sheetname
    c1.style = 13
    c1.y_axis.title = y_title
    c1.x_axis.title = x_title
    data = Reference(ws, min_col=2+start_row, min_row=1, max_col=1+start_row*2, max_row=15)
    c1.add_data(data, titles_from_data=True)
    titles = Reference(ws, min_col=1, min_row=2, max_row=15)
    c1.set_categories(titles)
    for i in range(3):
      s1 = c1.series[i]
      s1.graphicalProperties.line.solidFill = colors[i]
    ws.add_chart(c1, "A20")
  wb.save(url)

if __name__ == '__main__':
	### input varible based on your need 
    option = {
	'url': 'G:\\0508-马上消费\\duiwaijingmao\\5-对外经贸-全量\\test.csv',
	'lists' : ['other_var1','other_var2'],    ### besides nine scores which would add to columns
	'cus_type': 'other_var2',
	'y':'other_var1'
	}
    df = load_data(option)
    df['other_var2'].head()
    score_distribution(df,option['cus_type'],point)
    ks_cus_type(df,option['cus_type'],option['y'])
    cus_score(df,option['cus_type'],point)
    score_cus(df,option['cus_type'],point)
    insert_plot('score_cus.xlsx',5)
    insert_plot('cus_score.xlsx',9)

